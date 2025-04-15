from app.models.question import add_question as model_add_question
from datetime import datetime
from flask import Blueprint, render_template, session, request, flash, jsonify
import csv
import openpyxl
import re
from app.models.question import get_teacher_questions
from app.utils.auth import login_required
from app.utils.database import database
from app.models.teacher_services import (
    get_teacher_data, get_teacher_classes, create_exam, get_teacher_exams,
    update_exam_status, get_students_by_class_section, get_class_section_pairs,
    update_student_status, update_bulk_student_status, manage_submission,
    get_total_students, get_subject_name, calculate_completion_percentage
)
import google.generativeai as genai

teacher_bp = Blueprint('teacher', __name__, url_prefix='/teacher')
import os
from dotenv import load_dotenv

load_dotenv()


@teacher_bp.route('/')
@login_required(user_types=['teacher'])
def dashboard():
    return render_template('teacher/dashboard.html')


@teacher_bp.route('/generate_exam', methods=['GET', 'POST'])
@login_required(user_types=['teacher'])
def generate_exam():
    if request.method == 'POST':
        if not request.is_json:
            return {'error': 'Invalid request format'}, 400

        data = request.json
        user_id = session.get('user_id')
        user_name = session.get('name', 'Unknown Teacher')

        try:
            exam_id = create_exam(data, user_id, user_name)
            flash('Exam created successfully!', 'success')
            return jsonify({'success': True, 'exam_id': exam_id})
        except Exception as e:
            return {'error': str(e)}, 500

    # GET request - render the form
    return render_template('teacher/generate_exam.html', subjects=session['subjects'])


@teacher_bp.route('/api/subjects_by_class', methods=['GET'])
@login_required(user_types=['teacher'])
def get_subjects_by_class():
    class_id = request.args.get('class')
    if not class_id:
        return {'error': 'Class ID is required'}, 400

    try:
        user_email = session.get('user_id')
        teacher_data = get_teacher_data(user_email)

        if not teacher_data or 'classes_teached' not in teacher_data:
            return {'subjects': []}, 200

        classes_teached = teacher_data.get('classes_teached', {})
        subjects_for_class = []

        # Check if the class exists in classes_teached
        if class_id in classes_teached:
            # Get all sections for this class
            sections = classes_teached[class_id].get('sections', {})

            # Collect all subject IDs across all sections
            subject_ids = []
            for section, section_subjects in sections.items():
                subject_ids.extend(section_subjects)

            # Remove duplicates
            subject_ids = list(set(subject_ids))

            # Get subject details for each subject ID
            for subject_id in subject_ids:
                subject_name = get_subject_name(subject_id)
                subjects_for_class.append({
                    'id': subject_id,
                    'name': subject_name
                })

        return {'subjects': subjects_for_class}, 200
    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/api/questions', methods=['GET'])
@login_required(user_types=['teacher'])
def get_questions():
    subject_id = request.args.get('subject')
    class_id = int(request.args.get('class'))
    if not subject_id or not class_id:
        return {'error': 'Subject ID and Class ID are required'}, 400

    try:
        questions_ref = database.child('questions')
        all_questions = questions_ref.get()
        questions_list = []

        if all_questions:
            for q_id, q_data in all_questions.items():
                if q_data.get('subject_id') == subject_id and q_data.get('class') == class_id:
                    questions_list.append({
                        'id': q_id,
                        'text': q_data.get('text', ''),
                        'subject_id': subject_id,
                        'question_type': q_data.get('question_type', 'mcq'),
                        'difficulty': q_data.get('difficulty', 'medium'),
                        'created_by_name': q_data.get('created_by_name', 'Unknown Teacher'),
                        'created_at': q_data.get('created_at', datetime.now().strftime('%Y-%m-%d')),
                        'marks': q_data.get('marks', 1),
                        'options': q_data.get('options', {}),
                        'column_a': q_data.get('column_a', {}),
                        'column_b': q_data.get('column_b', {}),
                        'assertion': q_data.get('assertion', ''),
                        'reason': q_data.get('reason', ''),
                        'case_text': q_data.get('case_text', '')
                    })

        return {'questions': questions_list}
    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/manage_students')
@login_required(user_types=['teacher'])
def manage_students():
    return render_template('teacher/manage_students.html')


@teacher_bp.route('/api/exams_and_students', methods=['GET'])
@login_required(user_types=['teacher'])
def get_exams_and_students():
    try:
        teacher_email = session.get('user_id')
        if not teacher_email:
            return {'error': 'User not authenticated'}, 401

        # Get teacher's data and exams
        teacher_data = get_teacher_data(teacher_email)
        teacher_exams = get_teacher_exams(teacher_email)

        # Get student data for classes taught by this teacher
        class_section_pairs = get_class_section_pairs(teacher_data)
        student_data = get_students_by_class_section(class_section_pairs)

        return jsonify({
            'exams': teacher_exams,
            'students': student_data
        })

    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/api/update_student_status', methods=['POST'])
@login_required(user_types=['teacher'])
def api_update_student_status():
    try:
        data = request.json
        student_email = data.get('student_email')
        new_status = data.get('status')

        if not student_email or new_status is None:
            return {'error': 'Missing required fields'}, 400

        name = update_student_status(student_email, new_status)

        return jsonify({
            'success': True,
            'message': f"Student {name} status updated to {new_status}"
        })

    except ValueError as e:
        return {'error': str(e)}, 404
    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/api/update_bulk_status', methods=['POST'])
@login_required(user_types=['teacher'])
def api_update_bulk_status():
    try:
        data = request.json
        student_emails = data.get('student_emails', [])
        new_status = data.get('status')

        if not student_emails or new_status is None:
            return {'error': 'Missing required fields'}, 400

        updated_count = update_bulk_student_status(student_emails, new_status)

        return jsonify({
            'success': True,
            'message': f"Updated status to {new_status} for {updated_count} students"
        })

    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/api/add_submission_reason', methods=['POST'])
@login_required(user_types=['teacher'])
def add_submission_reason():
    try:
        data = request.json
        student_email = data.get('student_email')
        exam_id = data.get('exam_id')
        reason = data.get('reason')

        if not student_email or not exam_id or not reason:
            return {'error': 'Missing required fields'}, 400

        manage_submission(student_email, exam_id, reason)

        return jsonify({
            'success': True,
            'message': "Submission reason added for student"
        })

    except ValueError as e:
        return {'error': str(e)}, 404
    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/api/delete_submission', methods=['POST'])
@login_required(user_types=['teacher'])
def delete_submission():
    try:
        data = request.json
        student_email = data.get('student_email')
        exam_id = data.get('exam_id')

        if not student_email or not exam_id:
            return {'error': 'Missing required fields'}, 400

        success = manage_submission(student_email, exam_id, delete=True)

        if success:
            return jsonify({
                'success': True,
                'message': "Submission deleted successfully"
            })
        else:
            return jsonify({
                'success': False,
                'message': "No submission found for this student"
            })

    except ValueError as e:
        return {'error': str(e)}, 404
    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/api/toggle_exam_status', methods=['POST'])
@login_required(user_types=['teacher'])
def toggle_exam_status():
    try:
        data = request.json
        exam_id = data.get('exam_id')
        new_status = data.get('status')

        if exam_id is None or new_status is None:
            return {'error': 'Missing required fields'}, 400

        update_exam_status(exam_id, new_status)

        return jsonify({
            'success': True,
            'message': f"Exam status updated to {'enabled' if new_status else 'disabled'}"
        })

    except ValueError as e:
        return {'error': str(e)}, 403 if "Unauthorized" in str(e) else 404
    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/api/total_questions_by_teacher', methods=['GET'])
@login_required(user_types=['teacher'])
def questions_by_teacher():
    try:
        teacher_email = session.get('user_id')
        if not teacher_email:
            return {'error': 'User not authenticated'}, 401

        teacher_data = get_teacher_data(teacher_email)
        questions_total = teacher_data.get('questions_created', 0)

        return jsonify({'questions_created': questions_total})
    except Exception as e:
        return {'error': str(e)}, 500


@teacher_bp.route('/api/recent_activity', methods=['GET'])
@login_required(user_types=['teacher'])
def get_recent_activity():
    try:
        teacher_email = session.get('user_id')
        teacher_exams = get_teacher_exams(teacher_email)
        activity_items = []

        for exam_data in teacher_exams:
            # Calculate submissions percentage if any
            submissions = exam_data.get('submissions', {})
            submission_count = len(submissions) - (1 if ' ' in submissions else 0)  # Subtract empty entry

            activity_items.append({
                'type': 'exam_created',
                'title': f"{exam_data.get('title')}",
                'description': f"Created for Class {exam_data.get('class')} - {exam_data.get('subject_name')}",
                'status': 'Active' if exam_data.get('exam_status') else 'Inactive',
                'status_class': 'success' if exam_data.get('exam_status') else 'warning',
                'timestamp': exam_data.get('created_at'),
                'submission_count': submission_count
            })

        # Sort activities by timestamp (newest first)
        activity_items.sort(key=lambda x: datetime.strptime(x['timestamp'], '%Y-%m-%d %H:%M:%S'), reverse=True)

        # Limit to most recent 5 activities
        return jsonify({'activities': activity_items[:5]})

    except Exception as e:
        flash(f"Error in recent activity: {str(e)}")
        return jsonify({'error': str(e)}), 500


@teacher_bp.route('/api/class_performance', methods=['GET'])
@login_required(user_types=['teacher'])
def get_class_performance():
    try:
        teacher_email = session.get('user_id')
        teacher_data = get_teacher_data(teacher_email)
        classes_teached = get_teacher_classes(teacher_data)

        if not classes_teached:
            return jsonify({'class_performance': []})

        class_performance = []
        teacher_exams = get_teacher_exams(teacher_email)

        # Track submissions by class
        class_submissions = {}
        class_total_students = {}

        # First, calculate how many students are in each class
        for class_id, class_data in classes_teached.items():
            sections = class_data.get('sections', {})
            class_total_students[class_id] = 0

            # Count students in each section
            for section in sections:
                class_ref = database.child('classes').child(class_id).child(section).get()
                if class_ref:
                    # Subtract 1 for the null entry at index 0
                    student_count = len(class_ref) - 1 if isinstance(class_ref, list) else 0
                    class_total_students[class_id] += student_count

        # Then calculate submission rates for exams
        for exam_data in teacher_exams:
            exam_class = str(exam_data.get('class'))
            submissions = exam_data.get('submissions', {})

            # Remove the empty submission placeholder
            submission_count = len(submissions) - (1 if ' ' in submissions else 0)

            if exam_class not in class_submissions:
                class_submissions[exam_class] = {
                    'total_submissions': 0,
                    'expected_submissions': 0
                }

            class_submissions[exam_class]['total_submissions'] += submission_count

            # Expected submissions is the total students in that class
            expected = class_total_students.get(exam_class, 0)
            class_submissions[exam_class]['expected_submissions'] += expected

        # Calculate completion percentage for each class
        for class_id, data in class_submissions.items():
            if data['expected_submissions'] > 0:
                completion_percentage = (data['total_submissions'] / data['expected_submissions']) * 100

                # Determine color based on completion percentage
                color = 'success'
                if completion_percentage < 60:
                    color = 'warning'
                if completion_percentage < 40:
                    color = 'danger'

                class_performance.append({
                    'class': f"Class {class_id}",
                    'percentage': round(completion_percentage, 1),
                    'color': color
                })

        return jsonify({'class_performance': class_performance})

    except Exception as e:
        flash(f"Error in class performance: {str(e)}")
        return jsonify({'error': str(e)}), 500


@teacher_bp.route('/api/pending_submissions', methods=['GET'])
@login_required(user_types=['teacher'])
def get_pending_submissions():
    try:
        teacher_email = session.get('user_id')
        teacher_data = get_teacher_data(teacher_email)
        classes_teached = get_teacher_classes(teacher_data)

        if not classes_teached:
            return jsonify({'exam_submissions': {}})

        # Get all students in teacher's classes
        students_by_class = {}
        for class_id, class_data in classes_teached.items():
            sections = class_data.get('sections', {})
            students_by_class[class_id] = []
            for section in sections:
                # Get students in this section
                students_ref = database.child('students').get()
                if students_ref:
                    for email, student in students_ref.items():
                        if str(student.get('class')) == class_id and student.get('section') == section:
                            students_by_class[class_id].append({
                                'email': email.replace(',', '.'),  # Convert back from Firebase format
                                'name': student.get('name', 'Unknown Student'),
                                'roll_no': student.get('rollno', 'N/A')
                            })

        # Get active exams for each class taught by the teacher
        teacher_exams = get_teacher_exams(teacher_email)
        active_exams = []

        today = datetime.now().strftime('%Y-%m-%d')
        for exam_data in teacher_exams:
            if exam_data.get('exam_status'):
                exam_date = exam_data.get('exam_date')

                # Check if the exam is current or past due
                if exam_date and exam_date <= today:
                    active_exams.append({
                        'id': exam_data.get('id'),
                        'title': exam_data.get('title'),
                        'class': exam_data.get('class'),
                        'subject': exam_data.get('subject_name'),
                        'date': exam_date,
                        'submissions': exam_data.get('submissions', {})
                    })

        # Group pending submissions by exam
        exam_submissions = {}

        for exam in active_exams:
            exam_id = exam.get('id')
            exam_class = str(exam.get('class'))
            submissions = exam.get('submissions', {})

            submitted_emails = set()
            for key in submissions:
                submitted_emails.add(key.replace(',', '.'))

            # Find students who haven't submitted
            pending_students = []
            for student in students_by_class.get(exam_class, []):
                student_email = student.get('email')
                if student_email not in submitted_emails:
                    exam_date = datetime.strptime(exam.get('date'), '%Y-%m-%d')
                    days_overdue = (datetime.now() - exam_date).days

                    status = 'overdue' if days_overdue > 0 else 'due_soon' if days_overdue >= -2 else 'on_track'

                    pending_students.append({
                        'student_name': student.get('name'),
                        'student_email': student_email,
                        'days_overdue': days_overdue,
                        'status': status,
                        'initials': ''.join([name[0].upper() for name in student.get('name', 'US').split()[:2]])
                    })

            # Only add exams that have pending submissions
            if pending_students:
                # Sort students by days overdue (most overdue first)
                pending_students.sort(key=lambda x: x['days_overdue'], reverse=True)

                exam_submissions[exam_id] = {
                    'title': exam.get('title'),
                    'class': exam.get('class'),
                    'subject': exam.get('subject'),
                    'date': exam.get('date'),
                    'pending_students': pending_students
                }

        return jsonify({'exam_submissions': exam_submissions})

    except Exception as e:
        flash(f"Error in pending submissions: {str(e)}")
        return jsonify({'error': str(e)}), 500


@teacher_bp.route('/api/dashboard_stats', methods=['GET'])
@login_required(user_types=['teacher'])
def get_dashboard_stats():
    try:
        teacher_email = session.get('user_id')
        teacher_data = get_teacher_data(teacher_email)

        if not teacher_data:
            return jsonify({'error': 'Teacher data not found'}), 404

        # Get classes taught by this teacher
        classes_teached = get_teacher_classes(teacher_data)

        # 1. Count total students
        total_students = get_total_students(classes_teached)

        # 2. Count active exams
        teacher_exams = get_teacher_exams(teacher_email)
        active_exams = sum(1 for exam in teacher_exams if exam.get('exam_status'))

        # 3. Count questions in question bank
        total_questions = teacher_data.get('questions_created', 0)
        session['questions_created'] = total_questions
        # 4. Calculate average completion percentage
        avg_completion = calculate_completion_percentage(classes_teached, teacher_exams)

        return jsonify({
            'total_students': total_students,
            'active_exams': active_exams,
            'total_questions': total_questions,
            'avg_completion': avg_completion
        })

    except Exception as e:
        flash(f"Error in dashboard stats: {str(e)}")
        return jsonify({'error': str(e)}), 500


@teacher_bp.route('/add_questions_via', methods=['GET'])
@login_required(user_types=['teacher'])
def add_questions_via():
    teacher_questions = get_teacher_questions(session['user_id'])
    return render_template('teacher/add_questions_via.html', teacher_questions=teacher_questions)


@teacher_bp.route('/upload_spreadsheet', methods=['POST'])
@login_required(user_types=['teacher'])
def upload_spreadsheet():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400

    try:
        questions = []

        if file.filename.endswith('.csv'):
            # Process CSV with csv module
            csv_data = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(csv_data)

            for row in reader:
                question = process_question_row(row)
                questions.append(question)

        elif file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
            # Process Excel with openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # Get header row
            headers = [cell.value for cell in ws[1]]

            # Process each row after the header
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                question = process_question_row(row_dict)
                questions.append(question)

        else:
            return jsonify({'success': False, 'message': 'Unsupported file format'}), 400

        return jsonify({
            'success': True,
            'message': f'Successfully processed questions',
            'questions': questions
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error processing file: {str(e)}'}), 500


def process_question_row(row):
    """Helper function to process each question row"""


    question = {
        'question_type': row['Question Type'],
        'text': row['Question Text'],
        'difficulty': row['Difficulty'],
        'marks': int(row['Marks']),
        'created_by': session['user_id'],
        'created_by_type': 'teacher',
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'class': int(request.form.get('class', 0)),
        'subject_id': request.form.get('subject_id', '')
    }

    if row['Question Type'] == 'mcq':
        options = []
        for opt in ['Option A', 'Option B', 'Option C', 'Option D']:
            if opt in row and row[opt]:  # Check if option exists and is not empty
                options.append(row[opt])

        question['options'] = options
        question['correct_answer'] = row['Correct Answer']

    elif row['Question Type'] == 'true_false':
        question['options'] = ['True', 'False']
        question['correct_answer'] = row['Correct Answer']

    elif row['Question Type'] == 'fill_in_blanks':
        # Extract blanks from question text
        blanks = []
        for match in re.finditer(r'\[blank\]', row['Question Text']):
            blanks.append(row['Correct Answer'])
        question['blanks'] = blanks

    return question

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if GEMINI_API_KEY:
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        print("Gemini API configured successfully.")
    except Exception as config_err:
        print(f"Error configuring Gemini API: {config_err}")
        # Handle configuration error appropriately
else:
    print("Warning: GEMINI_API_KEY not found in environment variables. PDF generation will fail.")


# --- Helper Function for AI Prompt ---
def create_ai_prompt(class_name, subject_name, count, type_preference='mcq', source_type='PDF content'):
    """Creates a standardized prompt for Gemini question generation."""
    # (Keep the implementation of create_ai_prompt exactly as it was in the previous correct version)
    type_instruction = ""
    if type_preference == 'mcq':
        type_instruction = "Generate Multiple Choice Questions (MCQ)."
    elif type_preference == 'true_false':
        type_instruction = "Generate True/False questions."
    elif type_preference == 'fill_in_blanks':
        type_instruction = "Generate Fill in the Blanks questions."
    elif type_preference == 'mixed':
        type_instruction = "Generate a mix of question types (MCQ, True/False, Fill in the Blanks)."
    else:
        type_instruction = f"Generate {type_preference} questions."  # Default fallback

    output_format_instruction = f"""
**Output Format Requirements:**
- Provide EXACTLY {count} questions based on the provided {source_type}.
- Number each question sequentially starting from 1.
- Follow the specific format for EACH question type precisely:

*For Multiple Choice (MCQ):*
NUMBER. Question text?
a) Option A text
b) Option B text
c) Correct Option C text *
d) Option D text
(Use letters a, b, c, d. Mark the single correct answer CLEARLY with an asterisk * at the VERY END of the correct option line, AFTER the text.)

*For True/False:*
NUMBER. Statement text?
a) True *
b) False
(OR use 'a) True' and 'b) False *' if False is correct. Mark the correct one with *.)

*For Fill in the Blanks (FITB):*
NUMBER. Question text with one or more [blank] placeholders.
Answer: Correct word/phrase for the first blank
Answer: Correct word/phrase for the second blank (if applicable)
(Use exactly '[blank]' for placeholders. Provide one 'Answer:' line for EACH blank.)

- Ensure a SINGLE blank line separates each complete question block (question text, options/answers).
- Do NOT include any introductory text, concluding remarks, commentary, difficulty labels, or marks within the output. Only provide the formatted questions.
"""

    full_prompt = f"""
You are an expert question generator creating educational content.
Class: {class_name}
Subject: {subject_name}
Source Material: Content from the provided {source_type}.
Number of Questions to Generate: {count}
Task: {type_instruction} Critically analyze the source material to create relevant and accurate questions.

{output_format_instruction}

Generate the questions now based *only* on the provided file content:
"""
    return full_prompt.strip()


@teacher_bp.route('/generate_from_pdf', methods=['POST'])
@login_required(user_types=['teacher'])
def generate_from_pdf():
    print("--- Received request for /generate_from_pdf ---")

    if not GEMINI_API_KEY:
        print("Error: AI Generation not configured (GEMINI_API_KEY missing).")
        return jsonify({'success': False, 'message': 'AI Generation is not configured on the server.'}), 503

    if 'file' not in request.files:
        print("Error: No 'file' part in the request.")
        return jsonify({'success': False, 'message': 'No PDF file uploaded'}), 400

    file = request.files['file']
    filename = file.filename

    if filename == '':
        print("Error: No file selected (empty filename).")
        return jsonify({'success': False, 'message': 'No file selected'}), 400

    print(f"Received file: {filename}")

    # File extension check
    if not filename.lower().endswith('.pdf'):
        print(f"Error: Invalid file extension for {filename}. Only .pdf is allowed.")
        return jsonify({'success': False, 'message': 'Invalid file extension. Only .pdf is allowed.'}), 400
    print(f"File extension check passed for: {filename}")

    # Input validation
    selected_class = request.form.get('class')
    selected_subject_id = request.form.get('subject_id')
    question_count = request.form.get('question_count', '5')
    question_type_pref = request.form.get('question_type_pref', 'mcq')

    print(
        f"Form Data - Class: {selected_class}, Subject ID: {selected_subject_id}, Count: {question_count}, Type Pref: {question_type_pref}")

    if not all([selected_class, selected_subject_id, question_count]):
        print("Error: Missing required form fields (class, subject, or count).")
        return jsonify({'success': False, 'message': 'Missing class, subject, or question count.'}), 400

    try:
        count = int(question_count)
        if not 1 <= count <= 20:
            raise ValueError("Invalid question count.")
    except ValueError:
        print(f"Error: Invalid question count value: {question_count}")
        return jsonify({'success': False, 'message': 'Invalid number of questions specified (must be 1-20).'}), 400

    generated_text = None
    ai_error = None
    uploaded_file_object = None

    try:
        # Read file content into memory
        file_content = file.read()

        # Prepare for AI Call
        subject_name = get_subject_name(selected_subject_id)
        class_name = f"Class {selected_class}"

        prompt = create_ai_prompt(
            class_name=class_name,
            subject_name=subject_name,
            count=count,
            type_preference=question_type_pref,
            source_type='PDF document'
        )

        print("\n--- Sending Prompt to Gemini ---")
        print("-----------------------------\n")

        # Upload file to Gemini API directly from memory using BytesIO
        from io import BytesIO
        print(f"Uploading file to Gemini: {filename}")
        file_bytes = BytesIO(file_content)

        # Use the BytesIO object as the file source
        uploaded_file_object = genai.upload_file(
            path=file_bytes,  # Pass BytesIO directly instead of path
            display_name=filename,
            mime_type='application/pdf'
        )
        print(
            f"File uploaded successfully. Gemini File Name: {uploaded_file_object.name}, URI: {uploaded_file_object.uri}")

        # Generate Content using google-generativeai library
        model = genai.GenerativeModel(model_name="gemini-1.5-flash-latest")

        print("Generating content...")
        response = model.generate_content([prompt, uploaded_file_object])

        print("\n--- Processing Gemini Response ---")

        if response.parts:
            generated_text = response.text
            print("Successfully generated text from AI.")
        else:
            ai_error_reason = "Unknown reason"
            if response.prompt_feedback and response.prompt_feedback.block_reason:
                ai_error_reason = f"Blocked due to: {response.prompt_feedback.block_reason_message or response.prompt_feedback.block_reason}"
            elif not response.candidates:
                ai_error_reason = "No candidates generated (check prompt or model capabilities)"
            else:
                try:
                    finish_reason = response.candidates[0].finish_reason
                    ai_error_reason = f"Finished with reason: {finish_reason}"
                    if finish_reason != genai.types.Candidate.FinishReason.STOP:
                        ai_error_reason += " (Potential issue)"
                except (IndexError, AttributeError):
                    pass

            print(f"AI generation failed or produced no text. Reason: {ai_error_reason}")
            ai_error = f"AI could not generate questions. Reason: {ai_error_reason}"

    except genai.types.generation_types.StopCandidateException as e:
        print(f"AI Generation stopped unexpectedly: {e}")
        ai_error = f"AI generation stopped: {e}"
    except Exception as e:
        print(f"Error during AI processing or file handling: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        ai_error = f'Error processing PDF or generating questions: {str(e)}'

    finally:
        # Delete the file from Gemini API storage if it was uploaded
        if uploaded_file_object and hasattr(uploaded_file_object, 'name'):
            try:
                print(f"Attempting to delete uploaded file from Gemini: {uploaded_file_object.name}")
                genai.delete_file(name=uploaded_file_object.name)
                print(f"Successfully deleted Gemini file: {uploaded_file_object.name}")
            except Exception as delete_err:
                print(f"Warning: Failed to delete uploaded file {uploaded_file_object.name} from Gemini: {delete_err}")
        elif uploaded_file_object:
            print("Warning: uploaded_file_object exists but has no 'name' attribute for deletion.")

    # Return result to frontend
    if generated_text:
        print("--- Returning Success to Frontend ---")
        return jsonify({
            'success': True,
            'generated_text': generated_text
        })
    else:
        print(f"--- Returning Failure to Frontend: {ai_error} ---")
        return jsonify({
            'success': False,
            'message': ai_error or 'An unknown error occurred during AI generation.'
        }), 500


@teacher_bp.route('/save_imported_questions', methods=['POST'])
@login_required(user_types=['teacher'])
def save_imported_questions():
    """
    Receives a list of questions (parsed and potentially edited on frontend)
    and saves them to the database. Assumes correct_answer is normalized.
    """
    if not request.is_json:
        return jsonify({'success': False, 'message': 'Invalid request format, JSON expected'}), 400

    data = request.json
    questions_to_save = data.get('questions', [])
    user_id = session.get('user_id')
    user_type = session.get('user_type', 'teacher')

    if not questions_to_save:
        return jsonify({'success': False, 'message': 'No questions provided to save'}), 400
    if not user_id:
        return jsonify({'success': False, 'message': 'User session expired or invalid'}), 401

    added_count = 0
    errors = []
    print('hello')
    try:
        for question_data in questions_to_save:
            # Basic validation before saving
            if not question_data.get('text') or \
                    not question_data.get('question_type') or \
                    question_data.get('class') is None or \
                    not question_data.get('subject_id'):
                errors.append(f"Skipped incomplete question: {str(question_data.get('text', 'N/A'))[:30]}...")
                continue  # Skip incomplete questions

            # Ensure necessary context fields are set
            question_data['created_by'] = user_id
            question_data['created_by_type'] = user_type
            if 'created_at' not in question_data:
                question_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Frontend should have normalized 'correct_answer' already
            # Frontend should have ensured 'marks' and 'class' are numbers

            success = model_add_question(question_data)
            if success:
                added_count += 1
            else:
                errors.append(f"Failed database save for question: {question_data.get('text', 'N/A')[:50]}...")

        print('second')
        # Update teacher's count if successful adds and user is a teacher
        if added_count > 0 and user_type == 'teacher':
            try:
                teacher_ref = database.child('administrators').child('teachers').child(session['user_id'])

                def transaction_update(current_data):
                    if current_data is None:
                        current_data = {'questions_created': 0}  # Initialize if somehow missing
                    current_count = current_data.get('questions_created', 0)
                    # Ensure count is integer before adding
                    if not isinstance(current_count, int):
                        current_count = 0
                    current_data['questions_created'] = current_count + added_count
                    return current_data

                result = teacher_ref.transaction(transaction_update)  # Capture result for debugging if needed

                # Update session count
                session_current_count = session.get('questions_created', 0)
                if not isinstance(session_current_count, int): session_current_count = 0  # Sanitize session count too
                session['questions_created'] = session_current_count + added_count
                session.modified = True

            except Exception as count_update_error:
                errors.append("Questions saved, but failed to update teacher's total count. Please report this.")

        if errors:
            return jsonify({
                'success': added_count > 0,
                'message': f"Processed {len(questions_to_save)} questions. Saved: {added_count}. Errors: {len(errors)}.",
                'added_count': added_count,
                'errors': errors  # Send back specific errors
            }), 207 if added_count > 0 else 400  # Use 207 for partial success
        else:
            return jsonify({
                'success': True,
                'message': f'Successfully added {added_count} questions to your question bank!',
                'added_count': added_count
            })

    except Exception as e:
        print(f"Error saving imported questions: {e}")
        import traceback
        traceback.print_exc()  # Print full traceback for debugging
        return jsonify({'success': False, 'message': f'An internal server error occurred: {str(e)}'}), 500
